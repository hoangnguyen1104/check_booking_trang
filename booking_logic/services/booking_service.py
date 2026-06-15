from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from .stock_service import enrich_excel
from .stock_allocation import build_available_batches, allocate_from_batches
from ..reports import build_legacy_result_text
from ..utils import BOOKING_RESULT_FILE, get_product_code, parse_quantity


def format_booking_result_file(note_col: int):
    workbook = load_workbook(BOOKING_RESULT_FILE)
    sheet = workbook.active
    fulfillment_column_letter = sheet.cell(row=1, column=note_col).column_letter
    note_column_letter = sheet.cell(row=1, column=note_col + 1).column_letter

    sheet.column_dimensions[fulfillment_column_letter].width = 18
    sheet.column_dimensions[note_column_letter].width = 55

    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=note_col).alignment = Alignment(vertical="top")
        sheet.cell(row=row, column=note_col + 1).alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(BOOKING_RESULT_FILE)


def read_booking_totals(filepath: str) -> dict:
    product_map = {}
    df = pd.read_excel(filepath, header=None, dtype=object)

    for product_code, amount in df.iloc[1:, [0, 1]].itertuples(index=False):
        if pd.isna(product_code):
            continue
        product_map[product_code] = product_map.get(product_code, 0) + parse_quantity(amount)

    return product_map


def update_booking_fulfillment(booking_file: str, available_batches: dict[str, list[dict]]) -> None:
    booking_df = pd.read_excel(booking_file, header=None, dtype=object)
    fulfillment_col = booking_df.shape[1]
    note_col = fulfillment_col + 1
    booking_df[fulfillment_col] = None
    booking_df[note_col] = None
    booking_df.iloc[0, fulfillment_col] = "Số lượng đáp ứng"
    booking_df.iloc[0, note_col] = "Ghi chú chọn lô"

    for row_idx in range(1, len(booking_df)):
        product_code = get_product_code(booking_df.iloc[row_idx, 0])
        if not product_code:
            continue

        order_quantity = parse_quantity(booking_df.iloc[row_idx, 1])
        threshold_percent = float(booking_df.iloc[row_idx, 2])
        _, fulfilled_quantity, _, note = allocate_from_batches(
            product_code,
            order_quantity,
            available_batches,
            threshold_percent, 
        )

        booking_df.iloc[row_idx, fulfillment_col] = fulfilled_quantity
        booking_df.iloc[row_idx, note_col] = note

    booking_df.to_excel(BOOKING_RESULT_FILE, index=False, header=False)
    format_booking_result_file(note_col)


def build_result(stock_file: str, booking_file: str, ref_date: datetime) -> str:
    stock_df, batch_count = enrich_excel(stock_file, ref_date)
    available_batches = build_available_batches(stock_df)
    update_booking_fulfillment(booking_file, available_batches)
