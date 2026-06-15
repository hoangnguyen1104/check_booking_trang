from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from .stock import enrich_excel, build_available_batches, allocate_from_batches
from .utils import (
    BATCH_PATTERN,
    BOOKING_RESULT_FILE,
    EXTRA_COLUMNS,
    display_upload_name,
    get_product_code,
    parse_quantity,
)


def format_booking_result_file(note_col: int):
    workbook = load_workbook(BOOKING_RESULT_FILE)
    sheet = workbook.active
    fulfillment_column_letter = sheet.cell(row=1, column=note_col).column_letter
    note_column_letter = sheet.cell(row=1, column=note_col + 1).column_letter

    sheet.column_dimensions[fulfillment_column_letter].width = 18
    sheet.column_dimensions[note_column_letter].width = 55

    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=note_col).alignment = Alignment(vertical="top")
        sheet.cell(row=row, column=note_col + 1).alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(BOOKING_RESULT_FILE)


def update_booking_fulfillment(booking_file: str, available_batches: dict[str, list[dict]]) -> None:
    booking_df = pd.read_excel(booking_file, header=None, dtype=object)
    fulfillment_col = booking_df.shape[1]
    note_col = fulfillment_col + 1
    booking_df[fulfillment_col] = None
    booking_df[note_col] = None
    booking_df.iloc[0, fulfillment_col] = "Số lượng đáp ứng"
    booking_df.iloc[0, note_col] = "Ghi chú chọn lô"

    for row_idx in range(1, len(booking_df)):
        product_code = get_product_code(booking_df.iloc[row_idx, 0])
        if not product_code:
            continue

        order_quantity = parse_quantity(booking_df.iloc[row_idx, 1])
        _, fulfilled_quantity, _, note = allocate_from_batches(
            product_code,
            order_quantity,
            available_batches,
        )

        booking_df.iloc[row_idx, fulfillment_col] = fulfilled_quantity
        booking_df.iloc[row_idx, note_col] = note

    booking_df.to_excel(BOOKING_RESULT_FILE, index=False, header=False)
    format_booking_result_file(note_col)


def read_booking_totals(filepath: str) -> dict:
    product_map = {}
    df = pd.read_excel(filepath, header=None, dtype=object)

    for product_code, amount in df.iloc[1:, [0, 1]].itertuples(index=False):
        if pd.isna(product_code):
            continue
        product_map[product_code] = product_map.get(product_code, 0) + parse_quantity(amount)

    return product_map


def build_legacy_result_text(
    stock_file: str,
    booking_file: str,
    stock_df,
    batch_count: int,
    ref_date: datetime,
    threshold_percent: float,
) -> str:
    lines = [
        f"--- {display_upload_name(stock_file)} ---",
        f"Ngày tham chiếu: {ref_date.strftime('%d/%m/%y')}",
        f"Ngưỡng date: {threshold_percent:g}%",
        f"Đã ghi {batch_count} dòng mã lô",
        f"Cột mới: {', '.join(EXTRA_COLUMNS)}",
        "Mẫu 5 dòng mã lô đầu:",
    ]

    sample = stock_df[
        stock_df.iloc[:, 0].astype(str).str.contains(
            r"\(\d{2}/\d{2}/\d{2}-\d{2}/\d{2}/\d{2}\)",
            na=False,
            regex=True,
        )
    ].head(5)
    lines.append(sample.to_string())

    p_orders = read_booking_totals(booking_file)
    for key, sl in p_orders.items():
        lines.append(f"{key}: {sl}")

        for row_idx in range(len(stock_df)):
            value = stock_df.iloc[row_idx, 0]
            if str(key) not in str(value):
                continue

            lines.append(str(value))
            p_remain = 0
            p_date = False

            for i in range(row_idx + 1, len(stock_df)):
                batch_value = stock_df.iloc[i, 0]
                match = BATCH_PATTERN.search(str(batch_value).strip())
                if not match:
                    break

                row_values = stock_df.iloc[i].tolist()
                ratio = float(row_values[6] or 0)
                if ratio >= threshold_percent:
                    lines.append(f"{row_values[0]} Tồn kho: {row_values[1]}   Date: {row_values[6]}")
                    p_date = True
                    p_remain += parse_quantity(row_values[1])

            if not p_date:
                lines.append("Không có lô nào còn đủ date")
            if p_remain < sl:
                lines.append(f"Không còn đủ tồn kho {p_remain}")
            break

        lines.extend(["", ""])

    return "\n".join(lines)


def build_result(stock_file: str, booking_file: str, ref_date: datetime, threshold_percent: float = 50) -> str:
    stock_df, batch_count = enrich_excel(stock_file, ref_date)
    available_batches = build_available_batches(stock_df, threshold_percent)
    update_booking_fulfillment(booking_file, available_batches)
    return build_legacy_result_text(stock_file, booking_file, stock_df, batch_count, ref_date, threshold_percent)
