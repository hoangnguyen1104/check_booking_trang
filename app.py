from datetime import datetime
import re
from pathlib import Path

import pandas as pd
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from werkzeug.utils import secure_filename


app = Flask(__name__)

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_FOLDER = BASE_DIR / "uploads"
OUTPUT_FOLDER = BASE_DIR / "outputs"
RESULT_FILE = OUTPUT_FOLDER / "result.txt"
BOOKING_RESULT_FILE = OUTPUT_FOLDER / "Booking_result.xlsx"

UPLOAD_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)

BATCH_PATTERN = re.compile(r"(\S+)\((\d{2}/\d{2}/\d{2})-(\d{2}/\d{2}/\d{2})\)")

EXTRA_COLUMNS = [
    "Ngày SX",
    "HSD",
    "Số ngày trong hạn",
    "Số ngày còn hạn",
    "Tỷ lệ (%)",
]


def parse_quantity(value) -> int:
    if pd.isna(value):
        return 0
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0


def calc_expiry_info(manufacture_date: datetime, expiry_date: datetime, ref_date: datetime) -> dict:
    total_days = (expiry_date - manufacture_date).days
    remaining_days = (expiry_date - ref_date).days
    ratio_percent = round(remaining_days / total_days * 100, 2) if total_days > 0 else 0

    return {
        "total_days": total_days,
        "remaining_days": remaining_days,
        "ratio_percent": ratio_percent,
    }


def enrich_excel(filepath: str, ref_date: datetime) -> tuple[pd.DataFrame, int]:
    df = pd.read_excel(filepath, header=None, dtype=object)
    total_columns = 2 + len(EXTRA_COLUMNS)

    df.iloc[1, 1] = "HCM/Stock/MT"
    df.iloc[2, 1] = "Số lượng"

    while df.shape[1] < total_columns:
        df[df.shape[1]] = None

    for col_idx, header in enumerate(EXTRA_COLUMNS, start=2):
        df.iloc[2, col_idx] = header

    batch_count = 0
    for row_idx in range(len(df)):
        if row_idx >= 3:
            df.iloc[row_idx, 1] = parse_quantity(df.iloc[row_idx, 1])

        value = df.iloc[row_idx, 0]
        if pd.isna(value):
            continue

        match = BATCH_PATTERN.search(str(value).strip())
        if not match:
            continue

        manufacture_str = match.group(2)
        expiry_str = match.group(3)
        manufacture_date = datetime.strptime(manufacture_str, "%d/%m/%y")
        expiry_date = datetime.strptime(expiry_str, "%d/%m/%y")
        info = calc_expiry_info(manufacture_date, expiry_date, ref_date)

        df.iloc[row_idx, 2] = manufacture_str
        df.iloc[row_idx, 3] = expiry_str
        df.iloc[row_idx, 4] = info["total_days"]
        df.iloc[row_idx, 5] = info["remaining_days"]
        df.iloc[row_idx, 6] = info["ratio_percent"]
        batch_count += 1

    return df, batch_count


def get_product_code(value) -> str | None:
    if pd.isna(value):
        return None

    text = str(value).strip()
    bracket_match = re.search(r"\[(.*?)\]", text)
    if bracket_match:
        return bracket_match.group(1).strip()

    return text if text else None


def display_upload_name(filepath: str) -> str:
    filename = Path(filepath).name
    for prefix in ("stock_", "booking_"):
        if filename.startswith(prefix):
            return filename[len(prefix):]
    return filename


def build_available_batches(stock_df: pd.DataFrame) -> dict[str, list[dict]]:
    available_batches = {}

    for row_idx in range(len(stock_df)):
        product_code = get_product_code(stock_df.iloc[row_idx, 0])
        if not product_code or BATCH_PATTERN.search(str(stock_df.iloc[row_idx, 0]).strip()):
            continue

        available_batches.setdefault(product_code, [])

        for batch_idx in range(row_idx + 1, len(stock_df)):
            batch_value = stock_df.iloc[batch_idx, 0]
            match = BATCH_PATTERN.search(str(batch_value).strip())
            if not match:
                break

            row_values = stock_df.iloc[batch_idx].tolist()
            ratio = float(row_values[6] or 0)
            quantity = parse_quantity(row_values[1])

            if ratio >= 50 and quantity > 0:
                available_batches[product_code].append(
                    {
                        "batch": str(row_values[0]),
                        "quantity": quantity,
                        "ratio": ratio,
                        "row_order": batch_idx,
                    }
                )

    for batches in available_batches.values():
        batches.sort(key=lambda batch: (batch["ratio"], batch["row_order"]))

    return available_batches


def total_batch_quantity(batches: list[dict]) -> int:
    return sum(batch["quantity"] for batch in batches)


def allocate_from_batches(product_code: str, order_quantity: int, available_batches: dict[str, list[dict]]) -> tuple[int, int, int, str]:
    batches = available_batches.get(product_code, [])
    stock_before = total_batch_quantity(batches)
    remaining_order = order_quantity
    fulfilled_quantity = 0
    note_lines = []

    for batch in batches:
        if remaining_order <= 0:
            break
        if batch["quantity"] <= 0:
            continue

        picked_quantity = min(remaining_order, batch["quantity"])
        batch["quantity"] -= picked_quantity
        remaining_order -= picked_quantity
        fulfilled_quantity += picked_quantity
        note_lines.append(f"{batch['batch']}: {picked_quantity} (Date {batch['ratio']}%)")

    stock_after = total_batch_quantity(batches)
    return stock_before, fulfilled_quantity, stock_after, "\n".join(note_lines)


def format_booking_result_file(note_col: int):
    workbook = load_workbook(BOOKING_RESULT_FILE)
    sheet = workbook.active
    fulfillment_column_letter = sheet.cell(row=1, column=note_col).column_letter
    note_column_letter = sheet.cell(row=1, column=note_col + 1).column_letter

    sheet.column_dimensions[fulfillment_column_letter].width = 18
    sheet.column_dimensions[note_column_letter].width = 55

    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=note_col).alignment = Alignment(vertical="top")
        sheet.cell(row=row, column=note_col + 1).alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(BOOKING_RESULT_FILE)


def update_booking_fulfillment(booking_file: str, available_batches: dict[str, list[dict]]) -> None:
    booking_df = pd.read_excel(booking_file, header=None, dtype=object)
    fulfillment_col = booking_df.shape[1]
    note_col = fulfillment_col + 1
    booking_df[fulfillment_col] = None
    booking_df[note_col] = None
    booking_df.iloc[0, fulfillment_col] = "Số lượng đáp ứng"
    booking_df.iloc[0, note_col] = "Ghi chú chọn lô"

    for row_idx in range(1, len(booking_df)):
        product_code = get_product_code(booking_df.iloc[row_idx, 0])
        if not product_code:
            continue

        order_quantity = parse_quantity(booking_df.iloc[row_idx, 1])
        _, fulfilled_quantity, _, note = allocate_from_batches(
            product_code,
            order_quantity,
            available_batches,
        )

        booking_df.iloc[row_idx, fulfillment_col] = fulfilled_quantity
        booking_df.iloc[row_idx, note_col] = note

    booking_df.to_excel(BOOKING_RESULT_FILE, index=False, header=False)
    format_booking_result_file(note_col)


def read_booking_totals(filepath: str) -> dict:
    product_map = {}
    df = pd.read_excel(filepath, header=None, dtype=object)

    for product_code, amount in df.iloc[1:, [0, 1]].itertuples(index=False):
        if pd.isna(product_code):
            continue
        product_map[product_code] = product_map.get(product_code, 0) + parse_quantity(amount)

    return product_map


def build_legacy_result_text(stock_file: str, booking_file: str, stock_df: pd.DataFrame, batch_count: int, ref_date: datetime) -> str:
    lines = [
        f"--- {display_upload_name(stock_file)} ---",
        f"Ngày tham chiếu: {ref_date.strftime('%d/%m/%y')}",
        f"Đã ghi {batch_count} dòng mã lô",
        f"Cột mới: {', '.join(EXTRA_COLUMNS)}",
        "Mẫu 5 dòng mã lô đầu:",
    ]

    sample = stock_df[
        stock_df.iloc[:, 0].astype(str).str.contains(
            r"\(\d{2}/\d{2}/\d{2}-\d{2}/\d{2}/\d{2}\)",
            na=False,
            regex=True,
        )
    ].head(5)
    lines.append(sample.to_string())

    p_orders = read_booking_totals(booking_file)
    for key, sl in p_orders.items():
        lines.append(f"{key}: {sl}")

        for row_idx in range(len(stock_df)):
            value = stock_df.iloc[row_idx, 0]
            if str(key) not in str(value):
                continue

            lines.append(str(value))
            p_remain = 0
            p_date = False

            for i in range(row_idx + 1, len(stock_df)):
                batch_value = stock_df.iloc[i, 0]
                match = BATCH_PATTERN.search(str(batch_value).strip())
                if not match:
                    break

                row_values = stock_df.iloc[i].tolist()
                ratio = float(row_values[6] or 0)
                if ratio >= 50:
                    lines.append(f"{row_values[0]} Tồn kho: {row_values[1]}   Date: {row_values[6]}")
                    p_date = True
                    p_remain += parse_quantity(row_values[1])

            if not p_date:
                lines.append("Không có lô nào còn đủ date")
            if p_remain < sl:
                lines.append(f"Không còn đủ tồn kho {p_remain}")
            break

        lines.extend(["", ""])

    return "\n".join(lines)


def build_result(stock_file: str, booking_file: str, ref_date: datetime) -> str:
    stock_df, batch_count = enrich_excel(stock_file, ref_date)
    available_batches = build_available_batches(stock_df)
    update_booking_fulfillment(booking_file, available_batches)
    return build_legacy_result_text(stock_file, booking_file, stock_df, batch_count, ref_date)


def save_upload(file_storage, prefix: str) -> str:
    filename = secure_filename(file_storage.filename)
    if not filename:
        raise ValueError("Tên file không hợp lệ")

    saved_path = UPLOAD_FOLDER / f"{prefix}_{filename}"
    file_storage.save(saved_path)
    return str(saved_path)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def process():
    stock_file = request.files.get("stock_file")
    booking_file = request.files.get("booking_file")
    ref_date_value = request.form.get("ref_date")

    if not stock_file or not booking_file or not ref_date_value:
        return render_template("index.html", error="Vui lòng chọn đủ 2 file và nhập ngày tham chiếu.")

    try:
        ref_date = datetime.strptime(ref_date_value, "%Y-%m-%d")
        stock_path = save_upload(stock_file, "stock")
        booking_path = save_upload(booking_file, "booking")
        result_text = build_result(stock_path, booking_path, ref_date)
        RESULT_FILE.write_text(result_text, encoding="utf-8")
    except Exception as exc:
        return render_template("index.html", error=f"Lỗi xử lý: {exc}")

    return render_template("index.html", success=True, result_preview=result_text)


@app.route("/download")
def download():
    return send_file(RESULT_FILE, as_attachment=True, download_name="result.txt")


@app.route("/download-booking")
def download_booking():
    return send_file(BOOKING_RESULT_FILE, as_attachment=True, download_name="Booking_result.xlsx")


if __name__ == "__main__":
    app.run(debug=True, use_reloader=False)
